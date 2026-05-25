import io
from collections import defaultdict
from functools import wraps

from django.db.models import Count
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404

from core.models import (
    Horario, Sesion, FranjaHoraria, Titulacion, CursoAcademico,
    NivelCurso, Aula, Notificacion, AuditoriaHorario, Asignatura,
    AsignaturaTitulacion, DisponibilidadProfesor, AsignacionProfesor,
)
from login.models import Usuario, PerfilProfesor, PerfilAlumno
from . import audit as audit_log
from .forms import (
    HorarioCreateForm, WorkflowForm, DisponibilidadForm,
    SesionManualForm, FranjaHorariaForm, FiltroHorarioAlumnoForm,
    TitulacionForm, AsignaturaForm, ConfigGeneracionForm,
)
from .generator import generate, validate_horario

# ── Color palettes ────────────────────────────────────────────────────────────
# Dark-theme (UI): semi-transparent on dark backgrounds
_ASIG_COLORS = [
    {'bg': 'rgba(29,78,216,.15)',   'border': 'rgba(59,130,246,.45)',  'text': '#93c5fd'},
    {'bg': 'rgba(124,58,237,.15)',  'border': 'rgba(167,139,250,.45)', 'text': '#c4b5fd'},
    {'bg': 'rgba(5,150,105,.15)',   'border': 'rgba(52,211,153,.45)',  'text': '#6ee7b7'},
    {'bg': 'rgba(180,83,9,.15)',    'border': 'rgba(251,191,36,.45)',  'text': '#fcd34d'},
    {'bg': 'rgba(219,39,119,.15)',  'border': 'rgba(249,168,212,.45)', 'text': '#f9a8d4'},
    {'bg': 'rgba(8,145,178,.15)',   'border': 'rgba(103,232,249,.45)', 'text': '#67e8f9'},
    {'bg': 'rgba(234,88,12,.15)',   'border': 'rgba(253,186,116,.45)', 'text': '#fdba74'},
    {'bg': 'rgba(13,148,136,.15)',  'border': 'rgba(94,234,212,.45)',  'text': '#5eead4'},
    {'bg': 'rgba(225,29,72,.15)',   'border': 'rgba(253,164,175,.45)', 'text': '#fda4af'},
    {'bg': 'rgba(67,56,202,.15)',   'border': 'rgba(165,180,252,.45)', 'text': '#a5b4fc'},
]
# Print/PDF: opaque pastels on white background
_PRINT_COLORS = [
    {'bg': '#dbeafe', 'border': '#3b82f6', 'text': '#1e40af'},
    {'bg': '#ede9fe', 'border': '#7c3aed', 'text': '#5b21b6'},
    {'bg': '#d1fae5', 'border': '#10b981', 'text': '#065f46'},
    {'bg': '#fef3c7', 'border': '#f59e0b', 'text': '#92400e'},
    {'bg': '#fce7f3', 'border': '#ec4899', 'text': '#9d174d'},
    {'bg': '#cffafe', 'border': '#06b6d4', 'text': '#164e63'},
    {'bg': '#ffedd5', 'border': '#f97316', 'text': '#9a3412'},
    {'bg': '#ccfbf1', 'border': '#14b8a6', 'text': '#134e4a'},
    {'bg': '#ffe4e6', 'border': '#f43f5e', 'text': '#9f1239'},
    {'bg': '#e0e7ff', 'border': '#6366f1', 'text': '#3730a3'},
]

_DIA_MAP_1BASED = {1: 'Lunes', 2: 'Martes', 3: 'Miércoles', 4: 'Jueves', 5: 'Viernes'}


# ── Decoradores de acceso ────────────────────────────────────────────────────

def decano_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.rol != 'DECANO':
            messages.error(request, "Acceso restringido al Decanato.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapped


def profesor_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.rol not in ('DECANO', 'PROFESOR'):
            messages.error(request, "Acceso restringido al profesorado.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapped


# ── Home / router ────────────────────────────────────────────────────────────

def schedule_home(request):
    rol = request.user.rol
    if rol == 'DECANO':
        return redirect('schedule:horarios_list')
    if rol == 'PROFESOR':
        return redirect('schedule:mi_horario')
    if rol == 'ALUMNO':
        return redirect('schedule:mi_horario')
    return render(request, "core/schedule/home.html")


# ── RF-01: Gestión de horarios (Decano) ─────────────────────────────────────

@decano_required
def horarios_list(request):
    horarios = (
        Horario.objects.all()
        .select_related('titulacion', 'nivel', 'curso_academico', 'creado_por')
        .annotate(num_sesiones=Count('sesiones'))
        .order_by('-curso_academico__fecha_inicio', 'titulacion', 'nivel')
    )

    # Filtros por GET
    estado_filter = request.GET.get('estado', '')
    titulacion_filter = request.GET.get('titulacion', '')
    if estado_filter:
        horarios = horarios.filter(estado=estado_filter)
    if titulacion_filter:
        horarios = horarios.filter(titulacion_id=titulacion_filter)

    context = {
        'horarios':       horarios,
        'estado_filter':  estado_filter,
        'titulacion_filter': titulacion_filter,
        'titulaciones':   Titulacion.objects.filter(activa=True),
        'estados':        Horario.Estado.choices,
        'contadores': {
            'total':     Horario.objects.count(),
            'borrador':  Horario.objects.filter(estado='BORRADOR').count(),
            'revision':  Horario.objects.filter(estado='REVISION').count(),
            'aprobado':  Horario.objects.filter(estado='APROBADO').count(),
            'rechazado': Horario.objects.filter(estado='RECHAZADO').count(),
        },
    }
    return render(request, "core/schedule/list.html", context)


@decano_required
def horario_create(request):
    if request.method == 'POST':
        form = HorarioCreateForm(request.POST)
        if form.is_valid():
            horario = form.save(commit=False)
            horario.creado_por = request.user
            horario.estado = Horario.Estado.BORRADOR
            horario.save()
            audit_log.log(
                usuario=request.user,
                accion='CREACION',
                modelo='Horario',
                objeto_id=horario.pk,
                valor_nuevo={'titulacion': str(horario.titulacion), 'nivel': str(horario.nivel)},
                descripcion=f"Horario creado en estado Borrador",
            )
            messages.success(request, f"Horario creado. Revisa las restricciones antes de generar.")
            return redirect('schedule:horario_restricciones', pk=horario.pk)
    else:
        form = HorarioCreateForm()
    return render(request, "core/schedule/crear.html", {'form': form})


def horario_detalle(request, pk):
    horario = get_object_or_404(
        Horario.objects.select_related('titulacion', 'nivel', 'curso_academico', 'creado_por', 'aprobado_por'),
        pk=pk,
    )
    sesiones = list(
        Sesion.objects.filter(horario=horario)
        .select_related('asignatura', 'profesor__usuario', 'aula', 'franja')
        .order_by('franja__dia_semana', 'franja__hora_inicio')
    )
    errores_validacion = list(dict.fromkeys(validate_horario(horario)))

    from collections import defaultdict

    _DIA_MAP = {1: 'Lunes', 2: 'Martes', 3: 'Miércoles', 4: 'Jueves', 5: 'Viernes'}

    if sesiones:
        # Derive the grid exclusively from the schedule's own sessions:
        # rows = unique (hora_inicio, hora_fin) that appear in sessions,
        # cols = days 1-5 for which active franjas with those exact times exist.
        # This avoids showing franjas from other schedule configurations.
        session_slots = sorted(set(
            (s.franja.hora_inicio, s.franja.hora_fin) for s in sesiones
        ))
        hi_list = list({sk[0] for sk in session_slots})
        hf_list = list({sk[1] for sk in session_slots})

        all_franjas = list(
            FranjaHoraria.objects.filter(
                activa=True,
                hora_inicio__in=hi_list,
                hora_fin__in=hf_list,
                dia_semana__in=[1, 2, 3, 4, 5],
            ).order_by('dia_semana', 'hora_inicio')
        )
        slot_keys = session_slots
        dias_vals = sorted(set(f.dia_semana for f in all_franjas))
    else:
        # No sessions yet: load TARDE franjas with days 1-5 only.
        # Among those, keep only the duration that is most represented
        # so multiple duration sets don't pollute the empty grid.
        base_qs = FranjaHoraria.objects.filter(activa=True, dia_semana__in=[1, 2, 3, 4, 5])
        if not horario.nivel.es_ultimo or not horario.titulacion.permite_maniana:
            base_qs = base_qs.filter(turno='TARDE')
        candidates = list(base_qs.order_by('dia_semana', 'hora_inicio'))

        if candidates:
            from collections import Counter
            dur_count = Counter(
                (f.hora_fin.hour * 60 + f.hora_fin.minute)
                - (f.hora_inicio.hour * 60 + f.hora_inicio.minute)
                for f in candidates
            )
            target_dur = dur_count.most_common(1)[0][0]
            all_franjas = [
                f for f in candidates
                if (f.hora_fin.hour * 60 + f.hora_fin.minute)
                   - (f.hora_inicio.hour * 60 + f.hora_inicio.minute) == target_dur
            ]
        else:
            all_franjas = candidates

        slot_keys = sorted(set((f.hora_inicio, f.hora_fin) for f in all_franjas))
        dias_vals = sorted(set(f.dia_semana for f in all_franjas))

    # Mapa (hora_inicio, dia_semana) → FranjaHoraria for cell DnD targets
    franja_by_slot_dia = {}
    for f in all_franjas:
        franja_by_slot_dia[(f.hora_inicio, f.dia_semana)] = f

    # Day labels — always 1-5 convention
    dias = [(dv, _DIA_MAP[dv]) for dv in dias_vals if dv in _DIA_MAP]
    dias_vals = [dv for dv, _ in dias]

    # Cuadrícula sesiones: hora_inicio → dia_semana → [sesion]
    _sgrid = defaultdict(lambda: defaultdict(list))
    for s in sesiones:
        _sgrid[s.franja.hora_inicio][s.franja.dia_semana].append(s)

    filas = []
    for hi, hf in slot_keys:
        celdas = []
        for dv in dias_vals:
            franja_obj = franja_by_slot_dia.get((hi, dv))
            celdas.append({
                'franja':   franja_obj,
                'sesiones': _sgrid[hi].get(dv, []),
            })
        filas.append({'hi': hi, 'hf': hf, 'celdas': celdas})

    workflow_form = WorkflowForm() if request.user.rol == 'DECANO' else None

    # Panel de disponibilidad de profesores
    from core.models import DisponibilidadProfesor
    prof_ids = list({s.profesor_id for s in sesiones if s.profesor_id})
    disps_qs = (
        DisponibilidadProfesor.objects
        .filter(profesor_id__in=prof_ids, curso_academico=horario.curso_academico)
        .select_related('profesor__usuario')
        .order_by('dia_semana', 'hora_inicio')
    )
    _disp_map = defaultdict(list)
    for d in disps_qs:
        _disp_map[d.profesor_id].append(d)

    # Días con sesión asignada por profesor (en cualquier horario del mismo curso)
    _sesiones_prof_dia = defaultdict(set)  # profesor_id → {dia_semana, ...}
    for s in sesiones:
        if s.profesor_id:
            _sesiones_prof_dia[s.profesor_id].add(s.franja.dia_semana)

    _dias_label = {1: 'L', 2: 'M', 3: 'X', 4: 'J', 5: 'V', 6: 'S', 7: 'D'}
    profs_disponibilidad = []
    for s in sesiones:
        if not s.profesor_id:
            continue
        pid = s.profesor_id
        if any(p['id'] == pid for p in profs_disponibilidad):
            continue
        disps = _disp_map.get(pid, [])
        dias_con_clase = _sesiones_prof_dia.get(pid, set())
        # Agrupar disponibilidad por día
        dias_disp = {}
        for d in disps:
            dias_disp.setdefault(d.dia_semana, []).append(d)
        # Construir resumen por día de la cuadrícula
        dias_resumen = []
        for dv in dias_vals:
            slots_dia = dias_disp.get(dv, [])
            tiene_clase = dv in dias_con_clase
            bloqueado = tiene_clase or any(x.tipo == 'INDISPONIBLE' for x in slots_dia)
            preferente = not tiene_clase and any(x.tipo == 'PREFERENTE' for x in slots_dia)
            disponible = not tiene_clase and any(x.tipo == 'DISPONIBLE' for x in slots_dia)
            estado = 'bloqueado' if bloqueado else ('preferente' if preferente else ('disponible' if disponible else 'sin_datos'))
            dias_resumen.append({
                'dia': dv,
                'label': _dias_label.get(dv, str(dv)),
                'estado': estado,
                'slots': slots_dia,
                'tiene_clase': tiene_clase,
            })
        profs_disponibilidad.append({
            'id': pid,
            'profesor': s.profesor,
            'dias': dias_resumen,
            'tiene_datos': bool(disps) or bool(dias_con_clase),
        })

    context = {
        'horario':               horario,
        'sesiones':              sesiones,
        'filas':                 filas,
        'dias':                  dias,
        'errores_validacion':    errores_validacion,
        'workflow_form':         workflow_form,
        'puede_editar':          request.user.rol == 'DECANO' and horario.estado == 'BORRADOR',
        'profs_disponibilidad':  profs_disponibilidad,
    }
    if request.GET.get('partial') == 'grid':
        return render(request, "core/schedule/_grid_partial.html", context)
    return render(request, "core/schedule/detalle.html", context)


# ── RF-04: Workflow de estados ───────────────────────────────────────────────

@decano_required
def horario_workflow(request, pk):
    horario = get_object_or_404(Horario, pk=pk)
    if request.method != 'POST':
        return redirect('schedule:horario_detalle', pk=pk)

    form = WorkflowForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Datos del formulario incorrectos.")
        return redirect('schedule:horario_detalle', pk=pk)

    accion = form.cleaned_data['accion']
    motivo = form.cleaned_data.get('motivo', '')

    transiciones_validas = {
        'BORRADOR':  ['REVISION'],
        'REVISION':  ['APROBADO', 'RECHAZADO'],
        'RECHAZADO': ['BORRADOR'],
        'APROBADO':  [],
    }

    if accion not in transiciones_validas.get(horario.estado, []):
        messages.error(request, f"Transición no permitida: {horario.estado} → {accion}.")
        return redirect('schedule:horario_detalle', pk=pk)

    if accion == 'RECHAZADO' and not motivo:
        messages.error(request, "Debe indicar el motivo del rechazo.")
        return redirect('schedule:horario_detalle', pk=pk)

    # RF-02.1: Validar horas lectivas antes de aprobar
    if accion == 'APROBADO':
        errores = validate_horario(horario)
        if errores:
            messages.warning(request, f"Hay {len(errores)} advertencias. El horario se aprueba con incidencias.")

    estado_anterior = horario.estado
    horario.estado = accion
    if accion == 'RECHAZADO':
        horario.motivo_rechazo = motivo
    if accion == 'APROBADO':
        horario.aprobado_por = request.user
    horario.save()

    audit_log.log_workflow(request.user, horario, estado_anterior, accion, motivo)

    # Notificar a los profesores afectados si el horario es aprobado/rechazado
    if accion in ('APROBADO', 'RECHAZADO'):
        _notificar_profesores_horario(horario, accion, motivo, request.user)

    msg_map = {
        'REVISION': 'Horario enviado a revisión.',
        'APROBADO': 'Horario aprobado correctamente.',
        'RECHAZADO': 'Horario rechazado.',
        'BORRADOR': 'Horario devuelto a borrador.',
    }
    messages.success(request, msg_map.get(accion, 'Estado actualizado.'))
    return redirect('schedule:horario_detalle', pk=pk)


# ── RF-02: Generación automática ─────────────────────────────────────────────

def _preflight(horario):
    """Devuelve lista de dicts con el estado prevuelo de cada asignatura."""
    at_qs = (
        AsignaturaTitulacion.objects
        .filter(titulacion=horario.titulacion, nivel=horario.nivel)
        .select_related('asignatura')
        .order_by('asignatura__semestre', 'asignatura__codigo')
    )
    result = []
    for at in at_qs:
        asig = at.asignatura
        asig_obj = AsignacionProfesor.objects.filter(
            asignatura=asig,
            curso_academico=horario.curso_academico,
            tipo='TITULAR',
        ).select_related('profesor__usuario').first()
        result.append({
            'asignatura':  asig,
            'tiene_prof':  asig_obj is not None,
            'profesor':    asig_obj.profesor if asig_obj else None,
            'semestre':    asig.semestre,
        })
    return result


@decano_required
def horario_generar(request, pk):
    from datetime import datetime, timedelta, time as dtime

    horario = get_object_or_404(Horario, pk=pk)
    if horario.estado != 'BORRADOR':
        messages.error(request, "Solo se puede generar en estado Borrador.")
        return redirect('schedule:horario_detalle', pk=pk)

    preflight     = _preflight(horario)
    sin_prof      = [p for p in preflight if not p['tiene_prof']]
    aulas_n       = Aula.objects.filter(activa=True).count()
    sesiones_prev = Sesion.objects.filter(horario=horario).count()

    config_form = ConfigGeneracionForm(request.POST or None)

    if request.method == 'POST' and config_form.is_valid():
        dias      = [int(d) for d in config_form.cleaned_data['dias']]
        hi        = config_form.cleaned_data['hora_inicio']
        hf        = config_form.cleaned_data['hora_fin']
        dur_min   = int(config_form.cleaned_data['duracion_min'])

        # Build FranjaHoraria objects from config (get_or_create so they persist
        # and can be referenced by Sesion FK; activa=True marks them as active)
        franjas_generadas = []
        TURNO_MAP = {1: 'MANIANA', 2: 'MANIANA', 3: 'TARDE', 4: 'TARDE', 5: 'TARDE'}
        base = datetime.combine(datetime.today(), hi)
        end  = datetime.combine(datetime.today(), hf)
        slots = []
        cursor = base
        while cursor + timedelta(minutes=dur_min) <= end:
            slots.append((cursor.time(), (cursor + timedelta(minutes=dur_min)).time()))
            cursor += timedelta(minutes=dur_min)

        for dia in sorted(dias):
            for slot_ini, slot_fin in slots:
                turno = 'MANIANA' if slot_ini < dtime(14, 0) else 'TARDE'
                nombre = f"{'LMXJV'[dia-1]}-{slot_ini.strftime('%H%M')}"
                franja, _ = FranjaHoraria.objects.get_or_create(
                    dia_semana=dia,
                    hora_inicio=slot_ini,
                    hora_fin=slot_fin,
                    defaults={'nombre': nombre, 'turno': turno, 'activa': True},
                )
                franja.activa = True
                franja.save(update_fields=['activa'])
                franjas_generadas.append(franja)

        n_sesiones, errores = generate(horario, franjas_override=franjas_generadas)
        audit_log.log_generacion(request.user, horario, n_sesiones, errores)
        if errores:
            messages.warning(request, f"{len(errores)} advertencia(s) durante la generación.")
            Notificacion.objects.create(
                destinatario=request.user,
                tipo=Notificacion.Tipo.ADVERTENCIA,
                motivo=Notificacion.Motivo.CONFLICTO,
                titulo=f"Generación con advertencias — {horario}",
                mensaje=f"Se generaron {n_sesiones} sesiones con {len(errores)} advertencia(s). Revisa el log para más detalles.",
            )
        messages.success(request, f"Generación completada: {n_sesiones} sesiones asignadas.")
        Notificacion.objects.create(
            destinatario=request.user,
            tipo=Notificacion.Tipo.EXITO,
            motivo=Notificacion.Motivo.CAMBIO_HORARIO,
            titulo=f"Horario generado — {horario}",
            mensaje=f"Se asignaron {n_sesiones} sesiones correctamente.",
        )
        return redirect('schedule:horario_log', pk=pk)

    # Preview: calculate slots from form values for the summary panel
    preview_slots = 0
    if config_form.is_bound and not config_form.is_valid():
        pass  # show errors
    else:
        # default preview (15:00-21:00, 90min, L-V = 4 slots × 5 dias = 20)
        preview_slots = 20

    context = {
        'horario':       horario,
        'preflight':     preflight,
        'sin_prof':      sin_prof,
        'aulas_n':       aulas_n,
        'sesiones_prev': sesiones_prev,
        'puede_generar': aulas_n > 0 and len(preflight) > 0,
        'config_form':   config_form,
    }
    return render(request, "core/schedule/generar.html", context)


@decano_required
def horario_eliminar(request, pk):
    horario = get_object_or_404(Horario, pk=pk)
    if request.method != 'POST':
        return redirect('schedule:horario_detalle', pk=pk)
    nombre = str(horario)
    audit_log.log(
        request.user, 'BORRADO', 'Horario', pk,
        valor_anterior={'titulacion': str(horario.titulacion), 'nivel': str(horario.nivel), 'estado': horario.estado},
        descripcion=f'Horario eliminado: {nombre}',
    )
    horario.delete()
    messages.success(request, f'Horario «{nombre}» eliminado correctamente.')
    return redirect('schedule:horarios_list')


@decano_required
def horario_log(request, pk):
    horario = get_object_or_404(
        Horario.objects.select_related('titulacion', 'nivel', 'curso_academico'),
        pk=pk,
    )

    # Entradas de auditoría para este horario
    logs = (
        AuditoriaHorario.objects
        .filter(modelo_afectado='Horario', objeto_id=horario.pk)
        .select_related('usuario')
        .order_by('-timestamp')
    )

    # Análisis por asignatura: requerido vs. asignado
    sesiones_qs = Sesion.objects.filter(horario=horario).select_related('asignatura', 'profesor__usuario', 'aula', 'franja')
    sesiones_por_asig = {}
    for s in sesiones_qs:
        sesiones_por_asig.setdefault(s.asignatura_id, []).append(s)

    at_qs = (
        AsignaturaTitulacion.objects
        .filter(titulacion=horario.titulacion, nivel=horario.nivel)
        .select_related('asignatura')
        .order_by('asignatura__semestre', 'asignatura__codigo')
    )

    analisis = []
    for at in at_qs:
        asig     = at.asignatura
        sesiones = sesiones_por_asig.get(asig.id, [])
        tiene_prof = AsignacionProfesor.objects.filter(
            asignatura=asig,
            curso_academico=horario.curso_academico,
            tipo='TITULAR',
        ).exists()
        analisis.append({
            'asignatura':  asig,
            'requeridas':  asig.sesiones_semanales,
            'asignadas':   len(sesiones),
            'sesiones':    sesiones,
            'tiene_prof':  tiene_prof,
            'ok':          len(sesiones) >= asig.sesiones_semanales,
            'semestre':    asig.semestre,
        })

    n_ok   = sum(1 for a in analisis if a['ok'])
    n_warn = len(analisis) - n_ok

    context = {
        'horario':  horario,
        'logs':     logs,
        'analisis': analisis,
        'n_ok':     n_ok,
        'n_warn':   n_warn,
        'pct_ok':   round(n_ok / len(analisis) * 100) if analisis else 0,
        'sem_list': [1, 2],
    }
    return render(request, "core/schedule/log.html", context)


# ── RF-07: Edición manual de sesiones ────────────────────────────────────────

@decano_required
def sesion_eliminar(request, pk):
    sesion = get_object_or_404(Sesion, pk=pk)
    horario_pk = sesion.horario.pk
    if sesion.horario.estado != 'BORRADOR':
        messages.error(request, "Solo se pueden editar sesiones en horarios en estado Borrador.")
        return redirect('schedule:horario_detalle', pk=horario_pk)
    if request.method == 'POST':
        audit_log.log_sesion_delete(request.user, sesion.pk, f"Sesión de {sesion.asignatura} eliminada manualmente")
        sesion.delete()
        messages.success(request, "Sesión eliminada.")
    return redirect('schedule:horario_detalle', pk=horario_pk)


@decano_required
def sesion_crear(request, horario_pk):
    horario = get_object_or_404(Horario, pk=horario_pk)
    if horario.estado != 'BORRADOR':
        messages.error(request, "Solo se pueden añadir sesiones en horarios en estado Borrador.")
        return redirect('schedule:horario_detalle', pk=horario_pk)

    if request.method == 'POST':
        form = SesionManualForm(request.POST)
        if form.is_valid():
            sesion = form.save(commit=False)
            sesion.horario = horario
            sesion.save()
            audit_log.log_sesion_create(request.user, sesion)
            messages.success(request, "Sesión añadida correctamente.")
            return redirect('schedule:horario_detalle', pk=horario_pk)
    else:
        form = SesionManualForm()

    return render(request, "core/schedule/sesion_form.html", {'form': form, 'horario': horario})


# ── RF-07.2: Mover sesión (drag-and-drop) ────────────────────────────────────

def _bloqueo_disp_aplica(disp, semestre_asignatura: int) -> bool:
    """
    Regla T1/S2: una indisponibilidad exclusiva de T1 no aplica a sesiones
    de S2 porque la restricción expira antes de que comience el segundo semestre.
    """
    return disp.aplica_a_semestre(semestre_asignatura)


@decano_required
def sesion_mover(request, pk):
    """AJAX: mueve una sesión a otra franja horaria (usado por drag-and-drop)."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)

    sesion  = get_object_or_404(Sesion, pk=pk)
    horario = sesion.horario

    if horario.estado != 'BORRADOR':
        return JsonResponse({'ok': False, 'error': 'Solo se pueden editar sesiones en estado Borrador.'}, status=400)

    franja_id = request.POST.get('franja_id')
    if not franja_id:
        return JsonResponse({'ok': False, 'error': 'franja_id requerido.'}, status=400)

    franja = get_object_or_404(FranjaHoraria, pk=franja_id)
    sem    = sesion.asignatura.semestre

    # RD-05: conflicto de profesor — solo si hay otra sesión del MISMO semestre.
    # S1 y S2 transcurren en periodos distintos del año, por lo que un profesor
    # puede tener asignaturas de ambos semestres en la misma franja.
    if Sesion.objects.filter(
        horario__curso_academico=horario.curso_academico,
        profesor=sesion.profesor,
        franja=franja,
        asignatura__semestre=sem,
    ).exclude(pk=sesion.pk).exists():
        return JsonResponse({'ok': False, 'error': 'Conflicto de profesor: ya tiene clase en esa franja.'}, status=400)

    # RD-12: conflicto de aula
    if Sesion.objects.filter(
        horario__curso_academico=horario.curso_academico,
        aula=sesion.aula,
        franja=franja,
    ).exclude(pk=sesion.pk).exists():
        return JsonResponse({'ok': False, 'error': 'Conflicto de aula: ya está ocupada en esa franja.'}, status=400)

    # RD-08 estudiantes: misma franja, mismo semestre, mismo horario
    if Sesion.objects.filter(
        horario=horario,
        franja=franja,
        asignatura__semestre=sem,
    ).exclude(pk=sesion.pk).exists():
        return JsonResponse({'ok': False, 'error': f'Conflicto: ya hay una asignatura del semestre {sem} en esa franja.'}, status=400)

    # RD-14: misma asignatura, mismo día
    if Sesion.objects.filter(
        horario=horario,
        asignatura=sesion.asignatura,
        franja__dia_semana=franja.dia_semana,
    ).exclude(pk=sesion.pk).exists():
        return JsonResponse({'ok': False, 'error': 'Conflicto: la asignatura ya tiene sesión ese día.'}, status=400)

    old_franja = sesion.franja
    sesion.franja = franja
    sesion.save()

    audit_log.log(
        usuario=request.user,
        accion='EDICION',
        modelo='Sesion',
        objeto_id=sesion.pk,
        valor_anterior={'franja': str(old_franja)},
        valor_nuevo={'franja': str(franja)},
        descripcion=f"Sesión de {sesion.asignatura} movida por drag-and-drop.",
    )
    return JsonResponse({'ok': True})


# ── RF-02.0: Hoja de restricciones previa a la generación ─────────────────────

@decano_required
def horario_restricciones(request, pk):
    """Hoja de restricciones: muestra checks del sistema + CRUD restricciones personalizadas."""
    from core.models import AsignacionProfesor as AP, RestriccionPersonalizada
    from schedule.forms import RestriccionPersonalizadaForm

    horario = get_object_or_404(Horario, pk=pk)
    if horario.estado != 'BORRADOR':
        messages.error(request, "Solo se puede generar en estado Borrador.")
        return redirect('schedule:horario_detalle', pk=pk)

    # Añadir restricción personalizada (POST)
    if request.method == 'POST':
        form_restr = RestriccionPersonalizadaForm(request.POST, horario=horario)
        if form_restr.is_valid():
            r = form_restr.save(commit=False)
            r.horario = horario
            dia_raw = form_restr.cleaned_data.get('dia_semana')
            r.dia_semana = int(dia_raw) if dia_raw != '' and dia_raw is not None else None
            r.save()
            messages.success(request, "Restricción añadida correctamente.")
        else:
            messages.error(request, "Revisa los campos de la restricción.")
        return redirect('schedule:horario_restricciones', pk=pk)

    form_restr = RestriccionPersonalizadaForm(horario=horario)

    preflight  = _preflight(horario)
    sin_prof   = [p for p in preflight if not p['tiene_prof']]
    aulas_n    = Aula.objects.filter(activa=True).count()
    franjas_n  = FranjaHoraria.objects.filter(activa=True).count()

    profs_con_asig = (
        AP.objects.filter(curso_academico=horario.curso_academico, tipo='TITULAR')
        .values_list('profesor_id', flat=True).distinct()
    )
    total_profs    = profs_con_asig.count()
    profs_con_disp = (
        DisponibilidadProfesor.objects
        .filter(curso_academico=horario.curso_academico, profesor_id__in=profs_con_asig)
        .values_list('profesor_id', flat=True).distinct().count()
    )

    DIAS_LABEL = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes',
                  5: 'Viernes', 6: 'Sábado'}

    restricciones_sistema = [
        {'codigo': 'RD-01', 'titulo': 'Sesiones semanales por asignatura',
         'descripcion': 'Cada asignatura se imparte el número de sesiones configurado.',
         'ok': len(preflight) > 0,
         'detalle': f"{len(preflight)} asignatura(s) en el plan de estudios."},
        {'codigo': 'RD-05', 'titulo': 'Sin solapamiento de profesor',
         'descripcion': 'Ningún profesor puede impartir dos asignaturas en la misma franja.',
         'ok': True, 'detalle': 'Verificado automáticamente durante la generación.'},
        {'codigo': 'RD-07', 'titulo': 'Profesor titular por asignatura',
         'descripcion': 'Cada asignatura debe tener un único profesor titular asignado.',
         'ok': len(sin_prof) == 0,
         'detalle': f"{len(preflight) - len(sin_prof)}/{len(preflight)} asignaturas con titular." if preflight else "Sin asignaturas."},
        {'codigo': 'RD-08', 'titulo': 'Disponibilidad del profesor',
         'descripcion': 'El generador respeta las franjas bloqueadas y prioriza las preferidas.',
         'ok': profs_con_disp > 0,
         'detalle': f"{profs_con_disp}/{total_profs} profesores han registrado disponibilidad."},
        {'codigo': 'RD-08', 'titulo': 'Sin solapamiento de estudiantes',
         'descripcion': 'Los alumnos del mismo semestre no tienen dos asignaturas en la misma franja.',
         'ok': True, 'detalle': 'Verificado por semestre durante la generación.'},
        {'codigo': 'RD-12', 'titulo': 'Sin solapamiento de aula',
         'descripcion': 'Un aula no puede albergar dos clases simultáneamente.',
         'ok': aulas_n > 0,
         'detalle': f"{aulas_n} aula(s) activa(s) disponible(s)."},
        {'codigo': 'RD-14', 'titulo': 'Sin repetición diaria',
         'descripcion': 'Una misma asignatura no puede repetirse el mismo día.',
         'ok': True, 'detalle': 'Verificado automáticamente durante la generación.'},
        {'codigo': 'RD-17', 'titulo': 'Solo franjas activas',
         'descripcion': 'El horario solo usa franjas horarias marcadas como activas.',
         'ok': franjas_n > 0,
         'detalle': f"{franjas_n} franja(s) activa(s) definida(s)."},
        {'codigo': 'RD-18', 'titulo': 'Turno tarde para grados individuales',
         'descripcion': 'Los grados individuales se imparten en turno de tarde.',
         'ok': True,
         'detalle': f"Permite mañana: {'Sí' if horario.titulacion.permite_maniana else 'No'}."},
    ]

    restricciones_custom = list(
        RestriccionPersonalizada.objects
        .filter(horario=horario, activa=True)
        .select_related('franja', 'asignatura')
        .order_by('tipo', 'creado_en')
    )
    # Añadir etiqueta legible para la plantilla
    for r in restricciones_custom:
        parts = []
        if r.franja:
            parts.append(f"{r.franja.get_dia_semana_display()} {r.franja.hora_inicio:%H:%M}–{r.franja.hora_fin:%H:%M}")
        if r.asignatura:
            parts.append(r.asignatura.nombre)
        if r.dia_semana is not None:
            parts.append(DIAS_LABEL.get(r.dia_semana, str(r.dia_semana)))
        r.etiqueta = ' · '.join(parts) if parts else '—'

    puede_continuar = aulas_n > 0 and len(preflight) > 0

    context = {
        'horario':               horario,
        'restricciones_sistema': restricciones_sistema,
        'restricciones_custom':  restricciones_custom,
        'form_restr':            form_restr,
        'sin_prof':              sin_prof,
        'aulas_n':               aulas_n,
        'franjas_n':             franjas_n,
        'total_profs':           total_profs,
        'profs_con_disp':        profs_con_disp,
        'puede_continuar':       puede_continuar,
        'preflight':             preflight,
    }
    return render(request, 'core/schedule/restricciones.html', context)


@decano_required
def restriccion_eliminar(request, pk, rpk):
    from core.models import RestriccionPersonalizada
    r = get_object_or_404(RestriccionPersonalizada, pk=rpk, horario_id=pk)
    r.delete()
    messages.success(request, "Restricción eliminada.")
    return redirect('schedule:horario_restricciones', pk=pk)


@decano_required
def sesion_conflictos(request, pk):
    """AJAX: devuelve los franja_pk que generarían conflicto si se mueve esta sesión."""
    sesion  = get_object_or_404(Sesion, pk=pk)
    horario = sesion.horario
    franjas = FranjaHoraria.objects.filter(activa=True)
    sem     = sesion.asignatura.semestre

    conflictos = []
    for f in franjas:
        razon = None

        # RD-05: conflicto de sesión existente del profesor — solo mismo semestre
        if sesion.profesor_id and Sesion.objects.filter(
            horario__curso_academico=horario.curso_academico,
            profesor=sesion.profesor, franja=f,
            asignatura__semestre=sem,
        ).exclude(pk=sesion.pk).exists():
            razon = 'Profesor ocupado en esta franja'

        # RD-12: conflicto de aula
        if not razon and sesion.aula_id and Sesion.objects.filter(
            horario__curso_academico=horario.curso_academico,
            aula=sesion.aula, franja=f,
        ).exclude(pk=sesion.pk).exists():
            razon = 'Aula ocupada en esta franja'

        # RD-08 estudiantes: semestre
        if not razon and Sesion.objects.filter(
            horario=horario, franja=f,
            asignatura__semestre=sem,
        ).exclude(pk=sesion.pk).exists():
            razon = f'Ya hay otra asignatura de semestre {sem}'

        # RD-14: misma asignatura mismo día
        if not razon and Sesion.objects.filter(
            horario=horario,
            asignatura=sesion.asignatura,
            franja__dia_semana=f.dia_semana,
        ).exclude(pk=sesion.pk).exists():
            razon = 'Esta asignatura ya tiene sesión ese día'

        if razon:
            conflictos.append({'franja_pk': f.pk, 'razon': razon})

    return JsonResponse({'conflictos': conflictos})


# ── RF-08 / RF-08.1: Disponibilidad del profesor ─────────────────────────────

@profesor_required
def disponibilidad(request):
    if request.user.rol == 'DECANO':
        profesor_id = request.GET.get('profesor')
        if profesor_id:
            perfil = get_object_or_404(PerfilProfesor, pk=profesor_id)
        else:
            perfil = None
        profesores = PerfilProfesor.objects.select_related('usuario').all()
    else:
        try:
            perfil = request.user.perfil_profesor
        except PerfilProfesor.DoesNotExist:
            messages.error(request, "No tiene perfil de profesor configurado.")
            return redirect('dashboard')
        profesores = None

    disponibilidades = (
        DisponibilidadProfesor.objects.filter(profesor=perfil).order_by('dia_semana', 'hora_inicio')
        if perfil else []
    )

    if request.method == 'POST' and perfil:
        form = DisponibilidadForm(request.POST)
        if form.is_valid():
            d = form.save(commit=False)
            d.profesor = perfil
            d.save()
            # RF-08.2: queda registrado en fecha_modificacion (auto_now=True)
            messages.success(request, "Disponibilidad guardada.")
            return redirect(request.path + (f"?profesor={perfil.pk}" if request.user.rol == 'DECANO' else ''))
    else:
        form = DisponibilidadForm()

    franjas = FranjaHoraria.objects.filter(activa=True).order_by('dia_semana', 'hora_inicio')

    context = {
        'perfil':           perfil,
        'profesores':       profesores,
        'disponibilidades': disponibilidades,
        'form':             form,
        'franjas':          franjas,
        'dias':             DisponibilidadProfesor.DIAS_SEMANA,
    }
    return render(request, "core/schedule/disponibilidad.html", context)


@profesor_required
def disponibilidad_eliminar(request, pk):
    d = get_object_or_404(DisponibilidadProfesor, pk=pk)
    if request.user.rol == 'PROFESOR' and d.profesor != request.user.perfil_profesor:
        messages.error(request, "No puede eliminar disponibilidades de otro profesor.")
        return redirect('schedule:disponibilidad')
    if request.method == 'POST':
        d.delete()
        messages.success(request, "Franja eliminada.")
    return redirect('schedule:disponibilidad')


# ── RF-09 / RF-11: Vista personal de horario ─────────────────────────────────

def mi_horario(request):
    rol = request.user.rol

    if rol == 'PROFESOR':
        try:
            perfil = request.user.perfil_profesor
        except PerfilProfesor.DoesNotExist:
            return render(request, "core/schedule/mi_horario.html", {'sin_perfil': True})

        sesiones = list(
            Sesion.objects.filter(
                profesor=perfil,
                horario__estado='APROBADO',
            )
            .select_related(
                'asignatura', 'aula', 'franja',
                'horario__curso_academico', 'horario__titulacion', 'horario__nivel',
            )
            .order_by('franja__dia_semana', 'franja__hora_inicio')
        )
        _annotate_colors(sesiones)
        filas, dias = _build_grid_filas_dias(sesiones)

        # Agenda view: one entry per day (1-based)
        agenda = []
        for dia_num in range(1, 6):
            dia_sesiones = [s for s in sesiones if s.franja.dia_semana == dia_num]
            agenda.append({
                'dia_num': dia_num,
                'dia_nom': _DIA_MAP_1BASED[dia_num],
                'sesiones': dia_sesiones,
            })

        context = {
            'rol':     'PROFESOR',
            'perfil':  perfil,
            'sesiones': sesiones,
            'filas':   filas,
            'dias':    dias,
            'agenda':  agenda,
        }

    elif rol == 'ALUMNO':
        try:
            perfil = request.user.perfil_alumno
        except PerfilAlumno.DoesNotExist:
            return render(request, "core/schedule/mi_horario.html", {'sin_perfil': True})

        # Find the approved horario for this student's titulación/nivel/curso
        horario_obj = None
        if perfil.titulacion and perfil.nivel and perfil.curso_academico:
            horario_obj = (
                Horario.objects
                .filter(
                    titulacion=perfil.titulacion,
                    nivel=perfil.nivel,
                    curso_academico=perfil.curso_academico,
                    estado='APROBADO',
                )
                .select_related('titulacion', 'nivel', 'curso_academico')
                .first()
            )

        semestre_filter = request.GET.get('semestre', '')
        enrolled_ids = set(perfil.matriculas.values_list('pk', flat=True))

        if horario_obj:
            qs = (
                Sesion.objects.filter(horario=horario_obj)
                .select_related('asignatura', 'aula', 'franja', 'profesor__usuario')
                .order_by('franja__dia_semana', 'franja__hora_inicio')
            )
            if semestre_filter:
                qs = qs.filter(asignatura__semestre=int(semestre_filter))
            sesiones = list(qs)
        else:
            sesiones = []

        _annotate_colors(sesiones)
        filas, dias = _build_grid_filas_dias(sesiones)

        # RF-14: detect overlaps only within enrolled sessions
        sesiones_matriculadas = [s for s in sesiones if s.asignatura_id in enrolled_ids]
        conflictos = _detectar_conflictos_alumno(sesiones_matriculadas)

        context = {
            'rol':              'ALUMNO',
            'perfil':           perfil,
            'horario_obj':      horario_obj,
            'sesiones':         sesiones,
            'filas':            filas,
            'dias':             dias,
            'enrolled_ids':     list(enrolled_ids),
            'semestre_filter':  semestre_filter,
            'conflictos':       conflictos,
        }

    else:
        context = {'rol': rol}

    return render(request, "core/schedule/mi_horario.html", context)


# ── RF-03: Exportación de horarios (Decano) ───────────────────────────────────

@decano_required
def horario_export_pdf(request, pk):
    """Renderiza una página standalone print-ready; el navegador convierte a PDF."""
    horario = get_object_or_404(
        Horario.objects.select_related('titulacion', 'nivel', 'curso_academico', 'aprobado_por'),
        pk=pk,
    )
    sesiones = list(
        Sesion.objects.filter(horario=horario)
        .select_related('asignatura', 'aula', 'franja', 'profesor__usuario')
        .order_by('franja__dia_semana', 'franja__hora_inicio')
    )
    _annotate_colors(sesiones, palette=_PRINT_COLORS)
    filas, dias = _build_grid_filas_dias(sesiones)

    from datetime import date
    return render(request, 'core/schedule/horario_pdf.html', {
        'horario':          horario,
        'sesiones':         sesiones,
        'filas':            filas,
        'dias':             dias,
        'fecha_generacion': date.today(),
    })


@decano_required
def horario_export_excel(request, pk):
    """Genera un archivo .xlsx con la cuadrícula del horario."""
    horario = get_object_or_404(
        Horario.objects.select_related('titulacion', 'nivel', 'curso_academico'),
        pk=pk,
    )
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "openpyxl no está instalado. Ejecute: pip install openpyxl")
        return redirect('schedule:horario_detalle', pk=pk)

    sesiones = list(
        Sesion.objects.filter(horario=horario)
        .select_related('asignatura', 'aula', 'franja', 'profesor__usuario')
        .order_by('franja__dia_semana', 'franja__hora_inicio')
    )
    filas, dias = _build_grid_filas_dias(sesiones)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horario"

    n_cols = len(dias) + 1  # franja col + one per day

    # ── Row 1: título principal ──────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Horario · {horario.titulacion} {horario.nivel.anio}º — {horario.curso_academico}"
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='DC2626')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # ── Row 2: subtítulo ─────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub_cell = ws.cell(row=2, column=1)
    from datetime import date
    sub_cell.value = f"Estado: {horario.get_estado_display()} · Generado: {date.today().strftime('%d/%m/%Y')}"
    sub_cell.font = Font(name='Calibri', size=10, color='A1A1AA')
    sub_cell.fill = PatternFill('solid', fgColor='18181B')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Row 3: cabecera de columnas ──────────────────────────────────────
    HDR_FILL = PatternFill('solid', fgColor='27272A')
    HDR_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    CENTER   = Alignment(horizontal='center', vertical='center')
    thin     = Side(style='thin', color='3F3F46')
    BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=3, column=1, value='Franja').font = HDR_FONT
    ws.cell(row=3, column=1).fill = HDR_FILL
    ws.cell(row=3, column=1).alignment = CENTER
    ws.cell(row=3, column=1).border = BORDER

    for col_i, (_, dia_nom) in enumerate(dias):
        cell = ws.cell(row=3, column=col_i + 2, value=dia_nom)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[3].height = 22

    # ── Data rows ────────────────────────────────────────────────────────
    CELL_FILL  = PatternFill('solid', fgColor='1C1C1E')
    EMPTY_FILL = PatternFill('solid', fgColor='09090B')

    for row_i, fila in enumerate(filas):
        row_num = 4 + row_i
        ws.row_dimensions[row_num].height = 52

        franja_label = f"{fila['hi'].strftime('%H:%M')} – {fila['hf'].strftime('%H:%M')}"
        fc = ws.cell(row=row_num, column=1, value=franja_label)
        fc.font = Font(name='Calibri', bold=True, size=10, color='A1A1AA')
        fc.fill = PatternFill('solid', fgColor='27272A')
        fc.alignment = CENTER
        fc.border = BORDER

        for col_i, celda in enumerate(fila['celdas']):
            col_num = col_i + 2
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if celda['sesiones']:
                lines = []
                for s in celda['sesiones']:
                    lines.append(s.asignatura.nombre)
                    prof_name = s.profesor.usuario.get_full_name() if s.profesor else '—'
                    lines.append(f"  {s.aula.codigo} · {prof_name}")
                    if s.grupo:
                        lines.append(f"  Grupo: {s.grupo}")
                cell.value = "\n".join(lines)
                cell.font = Font(name='Calibri', size=9, color='E4E4E7')
                cell.fill = CELL_FILL
            else:
                cell.fill = EMPTY_FILL

    # ── Column widths ────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 18
    for i in range(len(dias)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 30

    # ── Output ───────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = (
        f"horario_{horario.titulacion.codigo}_{horario.nivel.anio}o"
        f"_{horario.curso_academico.nombre}.xlsx"
    ).replace(' ', '_')

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ── RF-10 / RF-12 / RNF-09: Notificaciones ───────────────────────────────────

def notificaciones(request):
    notifs = (
        Notificacion.objects.filter(destinatario=request.user)
        .order_by('-fecha_creacion')
    )
    no_leidas = notifs.filter(leida=False).count()
    return render(request, "core/schedule/notificaciones.html", {
        'notificaciones': notifs,
        'no_leidas':      no_leidas,
    })


def notificacion_leer(request, pk):
    notif = get_object_or_404(Notificacion, pk=pk, destinatario=request.user)
    if request.method == 'POST':
        notif.leida = True
        notif.save()
    return redirect('schedule:notificaciones')


def notificaciones_leer_todas(request):
    if request.method == 'POST':
        Notificacion.objects.filter(destinatario=request.user, leida=False).update(leida=True)
    return redirect('schedule:notificaciones')


def notificacion_eliminar(request, pk):
    notif = get_object_or_404(Notificacion, pk=pk, destinatario=request.user)
    if request.method == 'POST':
        notif.delete()
    return redirect('schedule:notificaciones')


def notificaciones_eliminar_todas(request):
    if request.method == 'POST':
        Notificacion.objects.filter(destinatario=request.user).delete()
    return redirect('schedule:notificaciones')


# ── RF-05: Configuración del marco temporal (Decano) ─────────────────────────

@decano_required
def config_view(request):
    franjas = FranjaHoraria.objects.all().order_by('dia_semana', 'hora_inicio')
    titulaciones = Titulacion.objects.all()
    aulas = Aula.objects.all().order_by('tipo', 'codigo')
    cursos = CursoAcademico.objects.all()

    if request.method == 'POST' and 'franja_submit' in request.POST:
        form_franja = FranjaHorariaForm(request.POST)
        if form_franja.is_valid():
            f = form_franja.save()
            audit_log.log(request.user, 'CREACION', 'FranjaHoraria', f.pk,
                          valor_nuevo={'nombre': f.nombre, 'dia': f.dia_semana})
            messages.success(request, "Franja horaria añadida.")
            return redirect('schedule:config')
    else:
        form_franja = FranjaHorariaForm()

    context = {
        'franjas':       franjas,
        'titulaciones':  titulaciones,
        'aulas':         aulas,
        'cursos':        cursos,
        'form_franja':   form_franja,
    }
    return render(request, "core/schedule/config.html", context)


@decano_required
def franja_toggle(request, pk):
    franja = get_object_or_404(FranjaHoraria, pk=pk)
    if request.method == 'POST':
        franja.activa = not franja.activa
        franja.save()
        estado = "activada" if franja.activa else "desactivada"
        messages.success(request, f"Franja {franja} {estado}.")
    return redirect('schedule:config')


@decano_required
def franja_eliminar(request, pk):
    franja = get_object_or_404(FranjaHoraria, pk=pk)
    if request.method == 'POST':
        audit_log.log(request.user, 'BORRADO', 'FranjaHoraria', pk,
                      valor_anterior={'nombre': franja.nombre})
        franja.delete()
        messages.success(request, "Franja eliminada.")
    return redirect('schedule:config')


# ── RNF-06 / RNF-14: Registro de auditoría ───────────────────────────────────

@decano_required
def audit_view(request):
    logs = (
        AuditoriaHorario.objects.all()
        .select_related('usuario')
        .order_by('-timestamp')[:200]
    )
    severidad_filter = request.GET.get('severidad', '')
    modelo_filter = request.GET.get('modelo', '')
    if severidad_filter:
        logs = AuditoriaHorario.objects.filter(severidad=severidad_filter).select_related('usuario').order_by('-timestamp')[:200]
    if modelo_filter:
        logs = AuditoriaHorario.objects.filter(modelo_afectado__icontains=modelo_filter).select_related('usuario').order_by('-timestamp')[:200]

    return render(request, "core/schedule/audit.html", {
        'logs':             logs,
        'severidad_filter': severidad_filter,
        'modelo_filter':    modelo_filter,
        'severidades':      AuditoriaHorario.Severidad.choices,
    })


# ── RF-03 / RF-03.1: Informes ─────────────────────────────────────────────────

@decano_required
def informes(request):
    titulacion_id = request.GET.get('titulacion', '')
    curso_id = request.GET.get('curso', '')

    horarios_aprobados = Horario.objects.filter(estado='APROBADO').select_related(
        'titulacion', 'nivel', 'curso_academico'
    )
    if titulacion_id:
        horarios_aprobados = horarios_aprobados.filter(titulacion_id=titulacion_id)
    if curso_id:
        horarios_aprobados = horarios_aprobados.filter(curso_academico_id=curso_id)

    # Estadísticas para el informe
    datos_informe = []
    for h in horarios_aprobados:
        sesiones = Sesion.objects.filter(horario=h).count()
        datos_informe.append({
            'horario':    h,
            'sesiones_n': sesiones,
            'horas_total': sesiones * 2,
        })

    context = {
        'datos_informe': datos_informe,
        'titulaciones':  Titulacion.objects.filter(activa=True),
        'cursos':        CursoAcademico.objects.all(),
        'titulacion_filter': titulacion_id,
        'curso_filter':  curso_id,
    }
    return render(request, "core/schedule/informes.html", context)


# ── Helpers internos ──────────────────────────────────────────────────────────

def _build_grid(sesiones):
    """Construye dict {dia_semana: {franja_id: [sesion]}} para el template."""
    grid = {}
    for s in sesiones:
        dia = s.franja.dia_semana
        fid = s.franja.id
        if dia not in grid:
            grid[dia] = {}
        if fid not in grid[dia]:
            grid[dia][fid] = []
        grid[dia][fid].append(s)
    return grid


def _build_grid_filas_dias(sesiones):
    """
    Build (filas, dias) from a list of Sesion objects (with .franja pre-selected).
    Uses 1-based dia_semana (1=Lunes..5=Viernes) matching the generator convention.
    Returns:
        filas: list of {'hi', 'hf', 'celdas': [{'sesiones': [...]}]}
        dias:  list of (dia_num, dia_nom)
    """
    if not sesiones:
        return [], []

    slot_keys = sorted({(s.franja.hora_inicio, s.franja.hora_fin) for s in sesiones})
    dias_vals = sorted({s.franja.dia_semana for s in sesiones})

    _sgrid = defaultdict(lambda: defaultdict(list))
    for s in sesiones:
        _sgrid[s.franja.hora_inicio][s.franja.dia_semana].append(s)

    dias = [(dv, _DIA_MAP_1BASED[dv]) for dv in dias_vals if dv in _DIA_MAP_1BASED]

    filas = []
    for hi, hf in slot_keys:
        celdas = [{'sesiones': _sgrid[hi].get(dv, [])} for dv, _ in dias]
        filas.append({'hi': hi, 'hf': hf, 'celdas': celdas})

    return filas, dias


def _annotate_colors(sesiones, palette=None):
    """Assigns a stable color dict to each session via s.color (in-place)."""
    if palette is None:
        palette = _ASIG_COLORS
    unique_ids = sorted({s.asignatura_id for s in sesiones})
    id_to_idx = {aid: i for i, aid in enumerate(unique_ids)}
    for s in sesiones:
        s.color = palette[id_to_idx.get(s.asignatura_id, 0) % len(palette)]


def _detectar_conflictos_alumno(sesiones):
    """
    RF-14: Detecta solapamientos en la carga lectiva del alumno.
    Devuelve lista de tuplas (sesion_a, sesion_b) que solapan.
    """
    conflictos = []
    lista = list(sesiones)
    for i, s1 in enumerate(lista):
        for s2 in lista[i+1:]:
            if s1.franja_id == s2.franja_id:
                conflictos.append((s1, s2))
    return conflictos


def _notificar_profesores_horario(horario, accion, motivo, actor):
    """Crea notificaciones in-app para los profesores del horario (RF-10.1, RF-12)."""
    motivo_notif = 'APROBACION' if accion == 'APROBADO' else 'RECHAZO'
    tipo_notif   = 'EXITO' if accion == 'APROBADO' else 'ADVERTENCIA'

    profesores_ids = (
        Sesion.objects.filter(horario=horario)
        .values_list('profesor__usuario_id', flat=True)
        .distinct()
    )
    titulo = (
        f"Horario {horario.titulacion} {horario.nivel.anio}º — {horario.get_estado_display()}"
    )
    mensaje = (
        f"El horario ha sido {horario.get_estado_display().lower()} por {actor.get_full_name() or actor.email}."
        + (f" Motivo: {motivo}" if motivo else "")
    )
    for uid in profesores_ids:
        Notificacion.objects.create(
            destinatario_id=uid,
            tipo=tipo_notif,
            motivo=motivo_notif,
            titulo=titulo,
            mensaje=mensaje,
        )


# ── Curriculum: Titulaciones y Asignaturas (solo Decano) ─────────────────────

@decano_required
def curriculum_view(request):
    titulaciones = (
        Titulacion.objects.all()
        .prefetch_related('niveles')
        .order_by('nombre')
    )
    context = {
        'titulaciones': titulaciones,
        'total_asig':   Asignatura.objects.count(),
        'total_tit':    titulaciones.count(),
    }
    return render(request, 'core/schedule/curriculum.html', context)


@decano_required
def titulacion_crear(request):
    if request.method == 'POST':
        form = TitulacionForm(request.POST)
        if form.is_valid():
            tit = form.save()
            # Auto-crear niveles 1..duracion_anios
            for anio in range(1, tit.duracion_anios + 1):
                NivelCurso.objects.get_or_create(
                    titulacion=tit,
                    anio=anio,
                    defaults={'es_ultimo': anio == tit.duracion_anios},
                )
            audit_log.log(
                usuario=request.user, accion='CREACION', modelo='Titulacion',
                objeto_id=tit.pk,
                valor_nuevo={'nombre': tit.nombre, 'codigo': tit.codigo},
                descripcion=f'Titulación creada: {tit.nombre}',
            )
            messages.success(request, f'Titulación «{tit.nombre}» creada con {tit.duracion_anios} niveles.')
            return redirect('schedule:curriculum')
    else:
        form = TitulacionForm()
    return render(request, 'core/schedule/titulacion_form.html', {'form': form, 'accion': 'Crear'})


@decano_required
def titulacion_editar(request, pk):
    tit = get_object_or_404(Titulacion, pk=pk)
    if request.method == 'POST':
        form = TitulacionForm(request.POST, instance=tit)
        if form.is_valid():
            tit = form.save()
            # Sincronizar niveles si cambia duración
            existing = set(tit.niveles.values_list('anio', flat=True))
            for anio in range(1, tit.duracion_anios + 1):
                if anio not in existing:
                    NivelCurso.objects.create(
                        titulacion=tit, anio=anio,
                        es_ultimo=(anio == tit.duracion_anios),
                    )
            # Marcar último año correctamente
            tit.niveles.all().update(es_ultimo=False)
            tit.niveles.filter(anio=tit.duracion_anios).update(es_ultimo=True)
            audit_log.log(
                usuario=request.user, accion='EDICION', modelo='Titulacion',
                objeto_id=tit.pk,
                valor_nuevo={'nombre': tit.nombre},
                descripcion=f'Titulación editada: {tit.nombre}',
            )
            messages.success(request, f'Titulación «{tit.nombre}» actualizada.')
            return redirect('schedule:curriculum')
    else:
        form = TitulacionForm(instance=tit)
    return render(request, 'core/schedule/titulacion_form.html', {
        'form': form, 'tit': tit, 'accion': 'Editar',
    })


@decano_required
def titulacion_eliminar(request, pk):
    tit = get_object_or_404(Titulacion, pk=pk)
    if request.method == 'POST':
        nombre = str(tit)
        audit_log.log(
            usuario=request.user, accion='BORRADO', modelo='Titulacion',
            objeto_id=pk, valor_anterior={'nombre': nombre},
            descripcion=f'Titulación eliminada: {nombre}', severidad='WARN',
        )
        tit.delete()
        messages.success(request, f'Titulación «{nombre}» eliminada.')
    return redirect('schedule:curriculum')


@decano_required
def titulacion_detalle(request, pk):
    tit = get_object_or_404(
        Titulacion.objects.prefetch_related('niveles'), pk=pk
    )
    # Agrupar asignaturas por año y semestre
    plan = {}
    for at in (
        AsignaturaTitulacion.objects
        .filter(titulacion=tit)
        .select_related('asignatura', 'nivel')
        .order_by('nivel__anio', 'asignatura__semestre', 'asignatura__codigo')
    ):
        anio = at.nivel.anio
        sem  = at.asignatura.semestre
        plan.setdefault(anio, {}).setdefault(sem, []).append(at.asignatura)

    return render(request, 'core/schedule/titulacion_detalle.html', {
        'tit':  tit,
        'plan': plan,
    })


@decano_required
def asignatura_crear(request, tit_pk):
    tit = get_object_or_404(Titulacion, pk=tit_pk)
    if request.method == 'POST':
        form = AsignaturaForm(request.POST, titulacion=tit)
        if form.is_valid():
            anio = int(form.cleaned_data['anio'])
            nivel = get_object_or_404(NivelCurso, titulacion=tit, anio=anio)
            asig = form.save()
            AsignaturaTitulacion.objects.get_or_create(
                asignatura=asig, titulacion=tit, nivel=nivel,
            )
            audit_log.log(
                usuario=request.user, accion='CREACION', modelo='Asignatura',
                objeto_id=asig.pk,
                valor_nuevo={'codigo': asig.codigo, 'nombre': asig.nombre},
                descripcion=f'Asignatura creada: {asig}',
            )
            messages.success(request, f'Asignatura «{asig.nombre}» creada y asignada a {anio}º.')
            return redirect('schedule:titulacion_detalle', pk=tit_pk)
    else:
        form = AsignaturaForm(titulacion=tit)

    niveles = tit.niveles.all()
    return render(request, 'core/schedule/asignatura_form.html', {
        'form':    form,
        'tit':     tit,
        'niveles': niveles,
        'accion':  'Crear',
    })


@decano_required
def asignatura_editar(request, pk):
    asig = get_object_or_404(Asignatura, pk=pk)
    # titulacion de referencia (la primera vinculación)
    at_ref = AsignaturaTitulacion.objects.filter(asignatura=asig).select_related('titulacion', 'nivel').first()
    tit    = at_ref.titulacion if at_ref else None

    if request.method == 'POST':
        form = AsignaturaForm(request.POST, instance=asig, titulacion=tit)
        if form.is_valid():
            asig = form.save()
            # Actualizar nivel si cambió el año
            if tit and at_ref:
                anio = int(form.cleaned_data['anio'])
                nivel = get_object_or_404(NivelCurso, titulacion=tit, anio=anio)
                if at_ref.nivel != nivel:
                    at_ref.nivel = nivel
                    at_ref.save()
            audit_log.log(
                usuario=request.user, accion='EDICION', modelo='Asignatura',
                objeto_id=asig.pk,
                valor_nuevo={'nombre': asig.nombre},
                descripcion=f'Asignatura editada: {asig}',
            )
            messages.success(request, f'Asignatura «{asig.nombre}» actualizada.')
            return redirect('schedule:titulacion_detalle', pk=tit.pk) if tit else redirect('schedule:curriculum')
    else:
        initial = {}
        if at_ref:
            initial = {'titulacion': at_ref.titulacion, 'anio': at_ref.nivel.anio}
        form = AsignaturaForm(instance=asig, titulacion=tit, initial=initial)

    return render(request, 'core/schedule/asignatura_form.html', {
        'form':    form,
        'tit':     tit,
        'niveles': tit.niveles.all() if tit else [],
        'asig':    asig,
        'accion':  'Editar',
    })


@decano_required
def asignatura_eliminar(request, pk):
    asig = get_object_or_404(Asignatura, pk=pk)
    at   = AsignaturaTitulacion.objects.filter(asignatura=asig).select_related('titulacion').first()
    tit_pk = at.titulacion.pk if at else None
    if request.method == 'POST':
        nombre = str(asig)
        audit_log.log(
            usuario=request.user, accion='BORRADO', modelo='Asignatura',
            objeto_id=pk, valor_anterior={'codigo': asig.codigo, 'nombre': asig.nombre},
            descripcion=f'Asignatura eliminada: {nombre}', severidad='WARN',
        )
        asig.delete()
        messages.success(request, f'Asignatura «{nombre}» eliminada.')
    return redirect('schedule:titulacion_detalle', pk=tit_pk) if tit_pk else redirect('schedule:curriculum')
